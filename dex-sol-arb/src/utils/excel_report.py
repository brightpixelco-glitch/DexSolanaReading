"""Generate Excel report from paper_trades.db."""
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..executor.paper import DB_PATH, TradeDB

EXCEL_PATH = Path(__file__).parent.parent.parent / "paper_results.xlsx"

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _style_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border


def _auto_width(ws, ncols):
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def generate_report(db_path: str = str(DB_PATH),
                    output_path: str = str(EXCEL_PATH)) -> str:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required — run: pip install openpyxl")

    tdb = TradeDB(db_path)
    sessions = tdb.get_session_summary()
    trades = tdb.get_trades_with_session()
    pairs = tdb.get_per_pair_pnl()

    wb = openpyxl.Workbook()

    # ──────────────────────── Sheet 1: Summary ────────────────────────
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells('A1:H1')
    ws['A1'] = f"DexSolana Paper Trade Report — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Session", "RPC Provider", "Total Trades", "Total PnL (SOL)",
               "Avg Profit %", "Wins", "Losses", "Best Trade", "Worst Trade"]
    _style_header(ws, headers, row=3)

    for i, s in enumerate(sessions, 4):
        vals = [
            s["label"], s["rpc_provider"], s["total_trades"],
            round(s["total_profit_sol"], 6), round(s["avg_profit_pct"], 2),
            s["wins"], s["losses"], round(s["best_trade"], 6), round(s["worst_trade"], 6),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            if col in (4, 8, 9):
                cell.number_format = '#,##0.000000'
            elif col == 5:
                cell.number_format = '#,##0.00'
    _auto_width(ws, len(headers))

    # ──────────────────────── Sheet 2: Per-Pair PnL ───────────────────
    ws2 = wb.create_sheet("Per-Pair PnL")
    headers2 = ["RPC Provider", "Token Mint", "Trades", "Total PnL (SOL)", "Avg Profit %"]
    _style_header(ws2, headers2)
    for i, r in enumerate(pairs, 2):
        vals = [
            r["rpc_provider"], r["token_mint"], r["trades"],
            round(r["total_profit_sol"], 6), round(r["avg_profit_pct"], 2),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.border = thin_border
            if col == 4:
                cell.number_format = '#,##0.000000'
            elif col == 5:
                cell.number_format = '#,##0.00'
    _auto_width(ws2, len(headers2))

    # ──────────────────────── Sheet 3: All Trades ─────────────────────
    ws3 = wb.create_sheet("All Trades")
    headers3 = ["Time", "Session", "RPC", "Strategy", "Token", "DEX",
                "Side", "Size (SOL)", "Entry", "Exit", "Profit (SOL)",
                "Profit %", "Confidence"]
    _style_header(ws3, headers3)
    for i, r in enumerate(trades, 2):
        vals = [
            r["timestamp"], r["session_label"], r["session_rpc"],
            r["strategy"], r["token_mint"], r["dex"], r["side"],
            r["trade_size_sol"], r["entry_price"], r["exit_price"],
            r["profit_sol"], r["profit_percent"], r["confidence"],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=col, value=v)
            cell.border = thin_border
            if col in (8, 9, 10, 11):
                cell.number_format = '#,##0.000000'
            elif col == 12:
                cell.number_format = '#,##0.00'
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers3))}{len(trades)+1}"
    _auto_width(ws3, len(headers3))

    wb.save(output_path)
    return output_path
